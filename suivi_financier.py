#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de génération d'un fichier Excel de suivi financier personnel
Créé pour: Suivi Salaire, Épargne et Investissements
Devise: MAD (Dirham marocain)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from datetime import datetime, timedelta
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

# Constantes de couleurs
BLEU_MARINE = "1a365d"
VERT = "38a169"
GRIS = "718096"
ROUGE = "e53e3e"
BLANC = "FFFFFF"
BLEU_CLAIR = "4299e1"
GRIS_CLAIR = "e2e8f0"

# Données financières
SALAIRE_NET = 14000
DEPENSES_FIXES = {
    "Loyer": 2250,
    "Électricité": 200,
    "Eau": 250,
    "Internet": 45,
    "Téléphone": 30,
    "Autre": 20
}
EPARGNE_PERSONNELLE = 2000
EPARGNE_FAMILLE = 2000
INVESTISSEMENT_MIN = 1000

def create_workbook():
    """Crée le classeur Excel avec tous les onglets"""
    wb = Workbook()

    # Supprimer la feuille par défaut
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Créer les onglets dans l'ordre souhaité
    create_revenus_sheet(wb)
    create_depenses_fixes_sheet(wb)
    create_depenses_variables_sheet(wb)
    create_epargne_sheet(wb)
    create_investissements_sheet(wb)
    create_patrimoine_sheet(wb)
    create_dashboard_sheet(wb)

    return wb

def apply_header_style(cell):
    """Applique le style d'en-tête"""
    cell.font = Font(bold=True, color=BLANC, size=12)
    cell.fill = PatternFill(start_color=BLEU_MARINE, end_color=BLEU_MARINE, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def apply_cell_style(cell, bg_color=None, bold=False, font_color="000000"):
    """Applique un style à une cellule"""
    cell.font = Font(bold=bold, color=font_color)
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def format_currency(ws, cell_range):
    """Formate une plage de cellules en devise MAD"""
    for row in ws[cell_range]:
        for cell in row:
            cell.number_format = '#,##0 "MAD"'

def create_revenus_sheet(wb):
    """Crée l'onglet Revenus"""
    ws = wb.create_sheet("Revenus", 0)
    ws.sheet_properties.tabColor = VERT

    # Titre
    ws['A1'] = "SUIVI DES REVENUS"
    ws['A1'].font = Font(bold=True, size=16, color=BLEU_MARINE)
    ws.merge_cells('A1:F1')

    # En-têtes
    headers = ['Mois', 'Année', 'Salaire Net', 'Primes/Bonus', 'Autres Revenus', 'TOTAL']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        apply_header_style(cell)

    # Données d'exemple (Janvier à Mars 2025)
    mois = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
            'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']

    for i, mois_nom in enumerate(mois, 4):
        ws[f'A{i}'] = mois_nom
        ws[f'B{i}'] = 2025
        ws[f'C{i}'] = SALAIRE_NET if i <= 6 else None  # Pré-rempli jusqu'à mars
        ws[f'D{i}'] = 0 if i <= 6 else None
        ws[f'E{i}'] = 0 if i <= 6 else None
        ws[f'F{i}'] = f'=C{i}+D{i}+E{i}'

        for col in range(1, 7):
            cell = ws.cell(row=i, column=col)
            apply_cell_style(cell, GRIS_CLAIR if i % 2 == 0 else None)

    # Total annuel
    ws['A16'] = "TOTAL ANNUEL"
    ws['A16'].font = Font(bold=True, color=BLEU_MARINE)
    for col in range(3, 7):
        cell = ws.cell(row=16, column=col)
        col_letter = get_column_letter(col)
        cell.value = f'=SUM({col_letter}4:{col_letter}15)'
        apply_cell_style(cell, BLEU_CLAIR, bold=True, font_color=BLANC)

    # Moyenne mensuelle
    ws['A17'] = "MOYENNE MENSUELLE"
    ws['A17'].font = Font(bold=True, color=BLEU_MARINE)
    for col in range(3, 7):
        cell = ws.cell(row=17, column=col)
        col_letter = get_column_letter(col)
        cell.value = f'=AVERAGE({col_letter}4:{col_letter}15)'
        apply_cell_style(cell, VERT, bold=True, font_color=BLANC)

    # Format monétaire
    for row in range(4, 18):
        for col in range(3, 7):
            ws.cell(row=row, column=col).number_format = '#,##0 "MAD"'

    # Ajuster les largeurs de colonnes
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

def create_depenses_fixes_sheet(wb):
    """Crée l'onglet Dépenses Fixes"""
    ws = wb.create_sheet("Dépenses Fixes")
    ws.sheet_properties.tabColor = ROUGE

    # Titre
    ws['A1'] = "DÉPENSES FIXES MENSUELLES"
    ws['A1'].font = Font(bold=True, size=16, color=BLEU_MARINE)
    ws.merge_cells('A1:E1')

    # En-têtes
    headers = ['Catégorie', 'Montant Prévu', 'Montant Réel', 'Statut', 'Écart']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        apply_header_style(cell)

    # Données des dépenses fixes
    row = 4
    for categorie, montant in DEPENSES_FIXES.items():
        ws[f'A{row}'] = categorie
        ws[f'B{row}'] = montant
        ws[f'C{row}'] = montant  # Pré-rempli
        ws[f'D{row}'] = "Payé"
        ws[f'E{row}'] = f'=C{row}-B{row}'

        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            apply_cell_style(cell, GRIS_CLAIR if row % 2 == 0 else None)

        row += 1

    # Total
    total_row = row
    ws[f'A{total_row}'] = "TOTAL"
    ws[f'B{total_row}'] = f'=SUM(B4:B{row-1})'
    ws[f'C{total_row}'] = f'=SUM(C4:C{row-1})'
    ws[f'E{total_row}'] = f'=SUM(E4:E{row-1})'

    for col in range(1, 6):
        cell = ws.cell(row=total_row, column=col)
        apply_cell_style(cell, BLEU_MARINE, bold=True, font_color=BLANC)

    # Format monétaire
    for r in range(4, total_row + 1):
        for col in [2, 3, 5]:
            ws.cell(row=r, column=col).number_format = '#,##0 "MAD"'

    # Historique mensuel
    ws['A12'] = "HISTORIQUE MENSUEL"
    ws['A12'].font = Font(bold=True, size=14, color=BLEU_MARINE)
    ws.merge_cells('A12:F12')

    # En-têtes historique
    headers_hist = ['Mois', 'Total Dépenses', 'Budget', 'Écart', '% du Salaire', 'Commentaires']
    for col, header in enumerate(headers_hist, 1):
        cell = ws.cell(row=13, column=col)
        cell.value = header
        apply_header_style(cell)

    # Données historique (3 mois d'exemple)
    mois = ['Janvier 2025', 'Février 2025', 'Mars 2025']
    for i, mois_nom in enumerate(mois, 14):
        ws[f'A{i}'] = mois_nom
        ws[f'B{i}'] = sum(DEPENSES_FIXES.values())
        ws[f'C{i}'] = sum(DEPENSES_FIXES.values())
        ws[f'D{i}'] = f'=B{i}-C{i}'
        ws[f'E{i}'] = f'=B{i}/{SALAIRE_NET}'
        ws[f'F{i}'] = ""

        for col in range(1, 7):
            cell = ws.cell(row=i, column=col)
            apply_cell_style(cell)

    # Format
    for row in range(14, 17):
        for col in [2, 3, 4]:
            ws.cell(row=row, column=col).number_format = '#,##0 "MAD"'
        ws.cell(row=row, column=5).number_format = '0.0%'

    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20

def create_depenses_variables_sheet(wb):
    """Crée l'onglet Dépenses Variables"""
    ws = wb.create_sheet("Dépenses Variables")
    ws.sheet_properties.tabColor = "ffa500"

    # Titre
    ws['A1'] = "DÉPENSES VARIABLES"
    ws['A1'].font = Font(bold=True, size=16, color=BLEU_MARINE)
    ws.merge_cells('A1:F1')

    ws['A2'] = f"Budget disponible: ~6 205 MAD/mois (Reste à vivre)"
    ws['A2'].font = Font(italic=True, color=GRIS)
    ws.merge_cells('A2:F2')

    # En-têtes
    headers = ['Catégorie', 'Budget Prévu', 'Dépensé', 'Écart', '% Budget', 'Statut']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        apply_header_style(cell)

    # Catégories de dépenses variables
    categories = {
        'Alimentation': 2500,
        'Transport': 800,
        'Loisirs': 600,
        'Santé': 300,
        'Vêtements': 200,
        'Imprévus': 500,
        'Autres': 1305
    }

    row = 5
    for categorie, budget in categories.items():
        ws[f'A{row}'] = categorie
        ws[f'B{row}'] = budget
        ws[f'C{row}'] = 0  # À remplir
        ws[f'D{row}'] = f'=B{row}-C{row}'
        ws[f'E{row}'] = f'=C{row}/B{row}'
        ws[f'F{row}'] = f'=IF(E{row}>1,"⚠️ Dépassement","✓ OK")'

        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            apply_cell_style(cell, GRIS_CLAIR if row % 2 == 0 else None)

        row += 1

    # Total
    total_row = row
    ws[f'A{total_row}'] = "TOTAL"
    ws[f'B{total_row}'] = f'=SUM(B5:B{row-1})'
    ws[f'C{total_row}'] = f'=SUM(C5:C{row-1})'
    ws[f'D{total_row}'] = f'=SUM(D5:D{row-1})'

    for col in range(1, 6):
        cell = ws.cell(row=total_row, column=col)
        apply_cell_style(cell, BLEU_MARINE, bold=True, font_color=BLANC)

    # Format
    for r in range(5, total_row + 1):
        for col in [2, 3, 4]:
            ws.cell(row=r, column=col).number_format = '#,##0 "MAD"'
        ws.cell(row=r, column=5).number_format = '0.0%'

    # Détails des dépenses
    ws['A15'] = "DÉTAILS DES DÉPENSES DU MOIS"
    ws['A15'].font = Font(bold=True, size=14, color=BLEU_MARINE)
    ws.merge_cells('A15:E15')

    headers_detail = ['Date', 'Catégorie', 'Description', 'Montant', 'Mode de paiement']
    for col, header in enumerate(headers_detail, 1):
        cell = ws.cell(row=16, column=col)
        cell.value = header
        apply_header_style(cell)

    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

def create_epargne_sheet(wb):
    """Crée l'onglet Épargne"""
    ws = wb.create_sheet("Épargne")
    ws.sheet_properties.tabColor = VERT

    # Titre
    ws['A1'] = "SUIVI DE L'ÉPARGNE"
    ws['A1'].font = Font(bold=True, size=16, color=BLEU_MARINE)
    ws.merge_cells('A1:F1')

    # Section Épargne Personnelle
    ws['A3'] = "💰 ÉPARGNE PERSONNELLE"
    ws['A3'].font = Font(bold=True, size=14, color=VERT)
    ws.merge_cells('A3:F3')

    ws['A4'] = f"Objectif: Fonds d'urgence de 6 mois = {6 * SALAIRE_NET:,} MAD"
    ws['A4'].font = Font(italic=True, color=GRIS)
    ws.merge_cells('A4:F4')

    # En-têtes épargne personnelle
    headers = ['Mois', 'Versement', 'Retraits', 'Net', 'Solde Cumulé', 'Progression Objectif']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        apply_header_style(cell)

    # Données épargne personnelle
    mois = ['Janvier 2025', 'Février 2025', 'Mars 2025']
    solde_cumul = 0
    for i, mois_nom in enumerate(mois, 6):
        solde_cumul += EPARGNE_PERSONNELLE
        ws[f'A{i}'] = mois_nom
        ws[f'B{i}'] = EPARGNE_PERSONNELLE
        ws[f'C{i}'] = 0
        ws[f'D{i}'] = f'=B{i}-C{i}'
        ws[f'E{i}'] = solde_cumul
        ws[f'F{i}'] = f'=E{i}/{6 * SALAIRE_NET}'

        for col in range(1, 7):
            cell = ws.cell(row=i, column=col)
            apply_cell_style(cell, GRIS_CLAIR if i % 2 == 0 else None)

    # Ajouter des lignes vides pour les mois futurs
    for i in range(9, 18):
        ws[f'A{i}'] = f'Mois {i-5}'
        ws[f'B{i}'] = EPARGNE_PERSONNELLE
        ws[f'C{i}'] = 0
        ws[f'D{i}'] = f'=B{i}-C{i}'
        if i == 9:
            ws[f'E{i}'] = f'=E8+D{i}'
        else:
            ws[f'E{i}'] = f'=E{i-1}+D{i}'
        ws[f'F{i}'] = f'=E{i}/{6 * SALAIRE_NET}'

        for col in range(1, 7):
            cell = ws.cell(row=i, column=col)
            apply_cell_style(cell)

    # Format
    for row in range(6, 18):
        for col in [2, 3, 4, 5]:
            ws.cell(row=row, column=col).number_format = '#,##0 "MAD"'
        ws.cell(row=row, column=6).number_format = '0.0%'

    # Section Épargne Famille
    ws['A20'] = "👨‍👩‍👧‍👦 ÉPARGNE FAMILLE / PARENTS"
    ws['A20'].font = Font(bold=True, size=14, color=VERT)
    ws.merge_cells('A20:E20')

    # En-têtes épargne famille
    headers_fam = ['Mois', 'Versement', 'Total Versé', 'Notes', 'Statut']
    for col, header in enumerate(headers_fam, 1):
        cell = ws.cell(row=21, column=col)
        cell.value = header
        apply_header_style(cell)

    # Données épargne famille
    total_verse = 0
    for i, mois_nom in enumerate(mois, 22):
        total_verse += EPARGNE_FAMILLE
        ws[f'A{i}'] = mois_nom
        ws[f'B{i}'] = EPARGNE_FAMILLE
        ws[f'C{i}'] = total_verse
        ws[f'D{i}'] = ""
        ws[f'E{i}'] = "Versé"

        for col in range(1, 6):
            cell = ws.cell(row=i, column=col)
            apply_cell_style(cell, GRIS_CLAIR if i % 2 == 0 else None)

    # Lignes futures
    for i in range(25, 34):
        ws[f'A{i}'] = f'Mois {i-21}'
        ws[f'B{i}'] = EPARGNE_FAMILLE
        if i == 25:
            ws[f'C{i}'] = f'=C24+B{i}'
        else:
            ws[f'C{i}'] = f'=C{i-1}+B{i}'
        ws[f'D{i}'] = ""
        ws[f'E{i}'] = ""

        for col in range(1, 6):
            cell = ws.cell(row=i, column=col)
            apply_cell_style(cell)

    # Format
    for row in range(22, 34):
        for col in [2, 3]:
            ws.cell(row=row, column=col).number_format = '#,##0 "MAD"'

    # Résumé
    ws['A36'] = "📊 RÉSUMÉ ÉPARGNE TOTALE"
    ws['A36'].font = Font(bold=True, size=14, color=BLEU_MARINE)
    ws.merge_cells('A36:D36')

    ws['A37'] = "Épargne mensuelle totale:"
    ws['B37'] = EPARGNE_PERSONNELLE + EPARGNE_FAMILLE
    ws['B37'].number_format = '#,##0 "MAD"'
    ws['A37'].font = Font(bold=True)
    ws['B37'].font = Font(bold=True)

    ws['A38'] = "Taux d'épargne (% du salaire):"
    ws['B38'] = f'={(EPARGNE_PERSONNELLE + EPARGNE_FAMILLE) / SALAIRE_NET}'
    ws['B38'].number_format = '0.0%'
    ws['A38'].font = Font(bold=True)
    ws['B38'].font = Font(bold=True)

    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20

def create_investissements_sheet(wb):
    """Crée l'onglet Investissements Bourse"""
    ws = wb.create_sheet("Investissements Bourse")
    ws.sheet_properties.tabColor = BLEU_CLAIR

    # Titre
    ws['A1'] = "📈 PORTEFEUILLE BOURSE"
    ws['A1'].font = Font(bold=True, size=16, color=BLEU_MARINE)
    ws.merge_cells('A1:H1')

    # Versements mensuels
    ws['A3'] = "💵 VERSEMENTS MENSUELS"
    ws['A3'].font = Font(bold=True, size=14, color=VERT)
    ws.merge_cells('A3:E3')

    headers_vers = ['Mois', 'Versement Prévu', 'Versement Effectué', 'Versement Extra', 'TOTAL']
    for col, header in enumerate(headers_vers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        apply_header_style(cell)

    # Données versements
    mois = ['Janvier 2025', 'Février 2025', 'Mars 2025']
    for i, mois_nom in enumerate(mois, 5):
        ws[f'A{i}'] = mois_nom
        ws[f'B{i}'] = INVESTISSEMENT_MIN
        ws[f'C{i}'] = INVESTISSEMENT_MIN
        ws[f'D{i}'] = 0
        ws[f'E{i}'] = f'=C{i}+D{i}'

        for col in range(1, 6):
            cell = ws.cell(row=i, column=col)
            apply_cell_style(cell, GRIS_CLAIR if i % 2 == 0 else None)

    # Total
    ws['A8'] = "TOTAL INVESTI"
    ws['B8'] = f'=SUM(B5:B7)'
    ws['C8'] = f'=SUM(C5:C7)'
    ws['D8'] = f'=SUM(D5:D7)'
    ws['E8'] = f'=SUM(E5:E7)'
    for col in range(1, 6):
        cell = ws.cell(row=8, column=col)
        apply_cell_style(cell, BLEU_MARINE, bold=True, font_color=BLANC)

    # Format
    for row in range(5, 9):
        for col in range(2, 6):
            ws.cell(row=row, column=col).number_format = '#,##0 "MAD"'

    # Portefeuille
    ws['A11'] = "📊 PORTEFEUILLE D'ACTIONS/ETF"
    ws['A11'].font = Font(bold=True, size=14, color=BLEU_MARINE)
    ws.merge_cells('A11:I11')

    headers_port = ['Date Achat', 'Ticker/Nom', 'Secteur', 'Quantité', 'Prix Achat',
                    'Valeur Achat', 'Prix Actuel', 'Valeur Actuelle', '+/- Value', '% Gain/Perte']
    for col, header in enumerate(headers_port, 1):
        cell = ws.cell(row=12, column=col)
        cell.value = header
        apply_header_style(cell)

    # Exemples de positions (à remplir par l'utilisateur)
    exemples = [
        ['15/01/2025', 'EXEMPLE ETF', 'Indices', 10, 100, '=D13*E13', 105, '=D13*G13', '=H13-F13', '=I13/F13'],
        ['20/01/2025', 'EXEMPLE ACTION', 'Tech', 5, 200, '=D14*E14', 210, '=D14*G14', '=H14-F14', '=I14/F14']
    ]

    for i, exemple in enumerate(exemples, 13):
        for col, value in enumerate(exemple, 1):
            ws.cell(row=i, column=col, value=value)
            apply_cell_style(ws.cell(row=i, column=col))

    # Total portefeuille
    ws['A15'] = "TOTAL PORTEFEUILLE"
    ws['F15'] = f'=SUM(F13:F14)'
    ws['H15'] = f'=SUM(H13:H14)'
    ws['I15'] = f'=SUM(I13:I14)'
    ws['J15'] = f'=I15/F15'

    for col in range(1, 11):
        cell = ws.cell(row=15, column=col)
        apply_cell_style(cell, VERT, bold=True, font_color=BLANC)

    # Format
    for row in range(13, 16):
        for col in [4]:
            ws.cell(row=row, column=col).number_format = '0'
        for col in [5, 6, 7, 8, 9]:
            ws.cell(row=row, column=col).number_format = '#,##0 "MAD"'
        ws.cell(row=row, column=10).number_format = '0.00%'

    # Performance
    ws['A18'] = "📈 PERFORMANCE DU PORTEFEUILLE"
    ws['A18'].font = Font(bold=True, size=14, color=BLEU_MARINE)
    ws.merge_cells('A18:D18')

    ws['A19'] = "Valeur totale investie:"
    ws['B19'] = '=F15'
    ws['B19'].number_format = '#,##0 "MAD"'

    ws['A20'] = "Valeur actuelle:"
    ws['B20'] = '=H15'
    ws['B20'].number_format = '#,##0 "MAD"'

    ws['A21'] = "Gain/Perte total:"
    ws['B21'] = '=I15'
    ws['B21'].number_format = '#,##0 "MAD"'

    ws['A22'] = "Performance (%):"
    ws['B22'] = '=J15'
    ws['B22'].number_format = '0.00%'

    for row in range(19, 23):
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True, size=12)

    # Dividendes
    ws['D19'] = "💰 DIVIDENDES REÇUS"
    ws['D19'].font = Font(bold=True, color=VERT)
    ws.merge_cells('D19:F19')

    headers_div = ['Date', 'Ticker', 'Montant']
    for col, header in enumerate(headers_div, 4):
        cell = ws.cell(row=20, column=col)
        cell.value = header
        apply_header_style(cell)

    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15

def create_patrimoine_sheet(wb):
    """Crée l'onglet Patrimoine Global"""
    ws = wb.create_sheet("Patrimoine Global")
    ws.sheet_properties.tabColor = "9b59b6"

    # Titre
    ws['A1'] = "💎 PATRIMOINE GLOBAL"
    ws['A1'].font = Font(bold=True, size=16, color=BLEU_MARINE)
    ws.merge_cells('A1:F1')

    # En-têtes
    headers = ['Mois', 'Épargne Personnelle', 'Épargne Famille', 'Portefeuille Bourse',
               'TOTAL PATRIMOINE', 'Variation']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        apply_header_style(cell)

    # Données (récapitulatif mensuel)
    mois = ['Janvier 2025', 'Février 2025', 'Mars 2025']

    # Liens vers les autres onglets
    for i, mois_nom in enumerate(mois, 4):
        ws[f'A{i}'] = mois_nom
        # Référence vers Épargne (onglet Épargne, colonne E)
        ws[f'B{i}'] = f"=Épargne!E{i+2}"
        # Référence vers Épargne Famille (onglet Épargne, colonne C section famille)
        ws[f'C{i}'] = f"=Épargne!C{i+18}"
        # Référence vers Investissements (valeur actuelle du portefeuille)
        ws[f'D{i}'] = f"='Investissements Bourse'!H15"
        # Total
        ws[f'E{i}'] = f'=B{i}+C{i}+D{i}'
        # Variation
        if i == 4:
            ws[f'F{i}'] = 0
        else:
            ws[f'F{i}'] = f'=E{i}-E{i-1}'

        for col in range(1, 7):
            cell = ws.cell(row=i, column=col)
            apply_cell_style(cell, GRIS_CLAIR if i % 2 == 0 else None)

    # Ajouter lignes futures
    for i in range(7, 16):
        ws[f'A{i}'] = f'Mois {i-3}'
        ws[f'B{i}'] = 0
        ws[f'C{i}'] = 0
        ws[f'D{i}'] = 0
        ws[f'E{i}'] = f'=B{i}+C{i}+D{i}'
        ws[f'F{i}'] = f'=E{i}-E{i-1}'

        for col in range(1, 7):
            cell = ws.cell(row=i, column=col)
            apply_cell_style(cell)

    # Format
    for row in range(4, 16):
        for col in range(2, 7):
            ws.cell(row=row, column=col).number_format = '#,##0 "MAD"'

    # Statistiques
    ws['A18'] = "📊 STATISTIQUES"
    ws['A18'].font = Font(bold=True, size=14, color=BLEU_MARINE)
    ws.merge_cells('A18:D18')

    stats = [
        ('Patrimoine actuel:', '=E6'),
        ('Croissance moyenne/mois:', '=AVERAGE(F4:F6)'),
        ('Objectif fin 2025:', '=E6+9*((B6-B4)/2+(C6-C4)/2+1000)'),
        ("Mois d'épargne d'urgence:", f'=E6/{SALAIRE_NET}')
    ]

    for i, (label, formule) in enumerate(stats, 19):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = formule
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'B{i}'].font = Font(bold=True, size=12, color=VERT)
        if i != 22:
            ws[f'B{i}'].number_format = '#,##0 "MAD"'
        else:
            ws[f'B{i}'].number_format = '0.0 "mois"'

    # Graphique d'évolution
    chart = LineChart()
    chart.title = "Évolution du Patrimoine"
    chart.style = 10
    chart.y_axis.title = "Montant (MAD)"
    chart.x_axis.title = "Mois"
    chart.height = 10
    chart.width = 20

    data = Reference(ws, min_col=5, min_row=3, max_row=15)
    cats = Reference(ws, min_col=1, min_row=4, max_row=15)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "A25")

    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15

def create_dashboard_sheet(wb):
    """Crée l'onglet Dashboard"""
    ws = wb.create_sheet("Dashboard", 0)  # Premier onglet
    ws.sheet_properties.tabColor = "FFD700"

    # Titre principal
    ws['A1'] = "📊 TABLEAU DE BORD FINANCIER"
    ws['A1'].font = Font(bold=True, size=20, color=BLEU_MARINE)
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    ws['A2'] = f"Mise à jour: {datetime.now().strftime('%d/%m/%Y')}"
    ws['A2'].font = Font(italic=True, color=GRIS)
    ws.merge_cells('A2:H2')
    ws['A2'].alignment = Alignment(horizontal="center")

    # KPIs principaux
    ws['A4'] = "💰 INDICATEURS CLÉS"
    ws['A4'].font = Font(bold=True, size=14, color=BLEU_MARINE)
    ws.merge_cells('A4:D4')

    kpis = [
        ('Salaire Mensuel', f'=Revenus!F4', BLEU_MARINE),
        ('Dépenses Fixes', f'="Dépenses Fixes"!B10', ROUGE),
        ('Épargne Totale/Mois', f'={EPARGNE_PERSONNELLE + EPARGNE_FAMILLE}', VERT),
        ('Investissement/Mois', f'={INVESTISSEMENT_MIN}', BLEU_CLAIR),
    ]

    row = 5
    for label, formule, couleur in kpis:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = formule
        ws[f'A{row}'].font = Font(bold=True, size=11)
        ws[f'B{row}'].font = Font(bold=True, size=14, color=couleur)
        ws[f'B{row}'].number_format = '#,##0 "MAD"'
        ws.merge_cells(f'B{row}:C{row}')
        row += 1

    # Répartition du salaire
    ws['A10'] = "📈 RÉPARTITION DU SALAIRE"
    ws['A10'].font = Font(bold=True, size=14, color=BLEU_MARINE)
    ws.merge_cells('A10:D10')

    repartition = [
        ('Dépenses Fixes', f'="Dépenses Fixes"!B10', f'="Dépenses Fixes"!B10/{SALAIRE_NET}'),
        ('Épargne Personnelle', EPARGNE_PERSONNELLE, f'={EPARGNE_PERSONNELLE}/{SALAIRE_NET}'),
        ('Épargne Famille', EPARGNE_FAMILLE, f'={EPARGNE_FAMILLE}/{SALAIRE_NET}'),
        ('Investissements', INVESTISSEMENT_MIN, f'={INVESTISSEMENT_MIN}/{SALAIRE_NET}'),
        ('Reste à Vivre', f'={SALAIRE_NET}-"Dépenses Fixes"!B10-{EPARGNE_PERSONNELLE}-{EPARGNE_FAMILLE}-{INVESTISSEMENT_MIN}',
         f'=({SALAIRE_NET}-"Dépenses Fixes"!B10-{EPARGNE_PERSONNELLE}-{EPARGNE_FAMILLE}-{INVESTISSEMENT_MIN})/{SALAIRE_NET}')
    ]

    headers_rep = ['Catégorie', 'Montant', '% du Salaire']
    for col, header in enumerate(headers_rep, 1):
        cell = ws.cell(row=11, column=col)
        cell.value = header
        apply_header_style(cell)

    row = 12
    for cat, montant, pourcentage in repartition:
        ws[f'A{row}'] = cat
        ws[f'B{row}'] = montant
        ws[f'C{row}'] = pourcentage

        apply_cell_style(ws[f'A{row}'])
        apply_cell_style(ws[f'B{row}'])
        apply_cell_style(ws[f'C{row}'])

        ws[f'B{row}'].number_format = '#,##0 "MAD"'
        ws[f'C{row}'].number_format = '0.0%'
        row += 1

    # Total
    ws['A17'] = "TOTAL"
    ws['B17'] = SALAIRE_NET
    ws['C17'] = 1.0
    for col in range(1, 4):
        cell = ws.cell(row=17, column=col)
        apply_cell_style(cell, BLEU_MARINE, bold=True, font_color=BLANC)
    ws['B17'].number_format = '#,##0 "MAD"'
    ws['C17'].number_format = '0.0%'

    # Graphique camembert
    pie = PieChart()
    pie.title = "Répartition du Salaire"
    labels = Reference(ws, min_col=1, min_row=12, max_row=16)
    data = Reference(ws, min_col=2, min_row=11, max_row=16)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 12
    pie.width = 15

    ws.add_chart(pie, "E5")

    # Patrimoine et objectifs
    ws['A20'] = "💎 PATRIMOINE & OBJECTIFS"
    ws['A20'].font = Font(bold=True, size=14, color=BLEU_MARINE)
    ws.merge_cells('A20:D20')

    patrimoine_data = [
        ('Patrimoine Total Actuel', '="Patrimoine Global"!E6'),
        ('Épargne Personnelle Cumulée', '=Épargne!E8'),
        ('Épargne Famille Totale', '=Épargne!C24'),
        ('Valeur Portefeuille Bourse', '="Investissements Bourse"!H15'),
        ('Performance Bourse', '="Investissements Bourse"!J15'),
    ]

    row = 21
    for label, formule in patrimoine_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = formule
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True, size=12, color=VERT)
        if 'Performance' in label:
            ws[f'B{row}'].number_format = '0.00%'
        else:
            ws[f'B{row}'].number_format = '#,##0 "MAD"'
        ws.merge_cells(f'B{row}:C{row}')
        row += 1

    # Taux d'épargne
    ws['A27'] = "🎯 TAUX D'ÉPARGNE"
    ws['A27'].font = Font(bold=True, size=14, color=BLEU_MARINE)
    ws.merge_cells('A27:D27')

    ws['A28'] = "Taux d'épargne actuel:"
    ws['B28'] = f'={(EPARGNE_PERSONNELLE + EPARGNE_FAMILLE)}/{SALAIRE_NET}'
    ws['B28'].number_format = '0.0%'
    ws['A28'].font = Font(bold=True)
    ws['B28'].font = Font(bold=True, size=16, color=VERT)
    ws.merge_cells('B28:C28')

    ws['A29'] = f"Objectif: 28.5% ✓ ATTEINT!"
    ws['A29'].font = Font(bold=True, color=VERT)
    ws.merge_cells('A29:C29')

    ws['A30'] = "Mois d'épargne d'urgence disponibles:"
    ws['B30'] = '=Épargne!E8/Revenus!C4'
    ws['B30'].number_format = '0.0 "mois"'
    ws['A30'].font = Font(bold=True)
    ws['B30'].font = Font(bold=True, size=16, color=BLEU_CLAIR)
    ws.merge_cells('B30:C30')

    # Notes et conseils
    ws['E20'] = "📝 NOTES & CONSEILS"
    ws['E20'].font = Font(bold=True, size=14, color=BLEU_MARINE)
    ws.merge_cells('E20:H20')

    notes = [
        "✓ Taux d'épargne excellent (28.5%)",
        "✓ Objectif: 6 mois d'épargne d'urgence",
        "• Diversifier les investissements",
        "• Suivre mensuellement les performances",
        "• Réajuster le budget si nécessaire"
    ]

    for i, note in enumerate(notes, 21):
        ws[f'E{i}'] = note
        ws[f'E{i}'].font = Font(size=10, color=GRIS)
        ws.merge_cells(f'E{i}:H{i}')

    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

def main():
    """Fonction principale"""
    print("🚀 Génération du fichier Excel de suivi financier...")
    print("=" * 60)

    # Créer le classeur
    wb = create_workbook()

    # Sauvegarder
    filename = "Suivi_Financier_Personnel_2025.xlsx"
    wb.save(filename)

    print(f"\n✅ Fichier créé avec succès: {filename}")
    print("\n📊 Onglets créés:")
    print("  1. Dashboard - Vue d'ensemble et KPIs")
    print("  2. Revenus - Suivi mensuel des revenus")
    print("  3. Dépenses Fixes - Charges mensuelles fixes")
    print("  4. Dépenses Variables - Budget et suivi des dépenses")
    print("  5. Épargne - Épargne personnelle et familiale")
    print("  6. Investissements Bourse - Portefeuille d'investissement")
    print("  7. Patrimoine Global - Vue consolidée du patrimoine")
    print("\n💡 Données pré-remplies pour Janvier-Mars 2025")
    print("=" * 60)
    print("\n🎯 Votre situation financière:")
    print(f"  • Salaire net: {SALAIRE_NET:,} MAD")
    print(f"  • Dépenses fixes: {sum(DEPENSES_FIXES.values()):,} MAD")
    print(f"  • Épargne totale: {EPARGNE_PERSONNELLE + EPARGNE_FAMILLE:,} MAD/mois")
    print(f"  • Investissements: {INVESTISSEMENT_MIN:,} MAD/mois minimum")
    print(f"  • Taux d'épargne: {((EPARGNE_PERSONNELLE + EPARGNE_FAMILLE)/SALAIRE_NET)*100:.1f}%")
    print("\n✨ Le fichier est prêt à l'emploi!")

if __name__ == "__main__":
    main()
